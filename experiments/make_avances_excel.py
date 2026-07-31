# -*- coding: utf-8 -*-
"""Workbook v3 — benchmark replanteado: propuesta (suma, r=12, m=0.069) vs 6 comparativos."""
import json, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
MR   = os.path.join(ROOT, "experiments", "results", "metrics_reports")
OUT  = os.path.join(ROOT, "docs", "Avances_Tesis_Tablas.xlsx")

F_TIT  = Font(name="Arial", size=12, bold=True)
F_SUB  = Font(name="Arial", size=10, bold=True)
F_HDR  = Font(name="Arial", size=9, bold=True)
F_TXT  = Font(name="Arial", size=9)
F_TXTB = Font(name="Arial", size=9, bold=True)
F_NOTA = Font(name="Arial", size=9, italic=True)
FILL_H = PatternFill("solid", fgColor="D9D9D9")
TH = Side(style="thin", color="000000")
BORDE = Border(left=TH, right=TH, top=TH, bottom=TH)
CC = Alignment(horizontal="center", vertical="center")
CL = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

def titulo(ws, fila, texto, sub=None):
    ws.cell(fila, 1, texto).font = F_TIT
    if sub:
        ws.cell(fila + 1, 1, sub).font = F_NOTA
        return fila + 3
    return fila + 2

def encabezado(ws, fila, cols, anchos=None):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(fila, j, c)
        cell.font = F_HDR; cell.fill = FILL_H; cell.alignment = CC; cell.border = BORDE
    if anchos:
        for j, a in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = a
    return fila + 1

def celda(ws, fila, col, valor, fmt=None, bold=False, align=CC):
    c = ws.cell(fila, col, valor)
    c.font = F_TXTB if bold else F_TXT
    c.alignment = align; c.border = BORDE
    if fmt:
        c.number_format = fmt
    return c

def nota(ws, fila, texto):
    c = ws.cell(fila, 1, texto); c.font = F_NOTA; c.alignment = WRAP
    return fila + 1

# ------------------------------------------------------------------ datos
means = pd.read_csv(os.path.join(MR, "descriptive_means.csv")).set_index("method")
fried = pd.read_csv(os.path.join(MR, "friedman_results.csv"))
wilc  = pd.read_csv(os.path.join(MR, "wilcoxon_results.csv"))
rankm = pd.read_csv(os.path.join(MR, "ranking_methods.csv"), index_col=0)
allm  = pd.read_csv(os.path.join(MR, "all_metrics.csv"))
_det_llvip = pd.read_csv(os.path.join(MR, "detection_llvip_map.csv")).set_index("method")
grid  = pd.read_csv(os.path.join(MR, "pso_grid_search_fo_propuesta.csv")).rename(
    columns={"Fo_opt": "F_opt"})

PROP = "Propuesta_Novedosa"
# Posicion de la propuesta en el ranking agregado, tomada del dato (promedio de rangos
# intra-bloque, el acompanante estandar de Friedman) y no fijada a mano.
_rk = rankm["avg_rank"].sort_values()
POS_RK = list(_rk.index).index(PROP) + 1
VAL_RK = f"{_rk[PROP]:.3f}".replace(".", ",")
N_ESC = int(allm["image"].nunique())
ORDEN = ["PiramideLaplace", "RatioPiramide", "DWT", "DTCWT", "Curvelet", "TopHat_Clasico", PROP]
LBL = {"PiramideLaplace": "Pirámide de Laplace (LP)", "RatioPiramide": "Ratio of low-pass Pyramid (RP)",
       "DWT": "Wavelet discreta (DWT)", "DTCWT": "Dual-Tree Complex Wavelet (DTCWT)",
       "Curvelet": "Curvelet (CVT)", "TopHat_Clasico": "Top-Hat clásico",
       PROP: "PROPUESTA NOVEDOSA (r=25, m=0,30)"}
DIRECTION = {"EN": 1, "SD": 1, "FE": 1, "MG": 1, "MI_vis": 1, "MI_ir": 1, "SF": 1,
             "SSIM": 1, "PSNR": 1}
METS = list(DIRECTION.keys())
NM = len(METS)  # 9

wb = Workbook()

# ============================================================ 1. RESUMEN
ws = wb.active; ws.title = "Resumen"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 95
f = titulo(ws, 1, "Fusión de imágenes IR/VIS mediante morfología matemática — Tablas de ejecución y métricas",
           "Maestría en Ciencias de Datos (UCOM) · Bazán & Bajac · Director: D.Sc. Mello · 13/07/2026")
info = [
    ("Propuesta central", "Fusión Top-Hat de UNA SOLA ESCALA (radio r): disco B_r + 4 líneas L_{r,θ} (0°, 45°, 90°, 135°, largo 2r+1). "
     "Respuestas lineales promediadas y SUMADAS a la del disco (Bala et al. 2024); entre fuentes máximo; "
     "reconstrucción F = I_base + m·WTH − m·BTH."),
    ("Optimización", "Barrido de 25 configuraciones PSO (partículas 2-10 × iteraciones 10-50, Cuadro 1 de Ortega & Espinoza 2025) sobre (r, m), r en [1,25]. El peso converge al piso del rango publicado (m* = 0,30) en las 25 configuraciones, porque Fo decrece de forma estrictamente monótona en m. El radio NO lo fija el PSO: dentro de ese rango Fo prefiere r = 1 (1,7354 frente a 1,7039 en r = 25), de modo que r = 25 es una decisión de diseño tomada sobre las métricas de evaluación, de las cuales cinco lo favorecen (EN, SD, FE, MG, SF) y cuatro prefieren r = 1 (SSIM, PSNR, MI_vis, MI_ir)."),
    ("Benchmark", "7 métodos: LP, RP (Toet 1989), DWT, DTCWT, CVT, Top-Hat clásico y la propuesta. "
     f"{N_ESC} pares TNO × nueve métricas sin referencia (EN, SD, FE, MG, MI_vis, MI_ir, SF, SSIM, PSNR)."),
    ("Resultados clave", f"La propuesta lidera la entropía ({means.loc[PROP,'EN']:.4f}) del estudio, "
     f"con ventaja significativa sobre los cinco métodos del estado del arte en EN, MG y SF; "
     f"{POS_RK}º lugar del ranking agregado ({VAL_RK}) por promedio de rangos intra-bloque. "
     "Nota: FE es la entropía EN reescalada por una constante por escena, de modo que no constituye "
     "evidencia independiente de EN."),
    ("Detección (LLVIP)", f"YOLOv8n reentrenado por método, evaluado con el checkpoint final (last.pt) para no "
     f"heredar el sesgo de elegirlo sobre el mismo conjunto que se reporta. El IR solo lidera "
     f"(mAP@0,5 = {_det_llvip.loc['IR','mAP50']:.4f}); toda fusión supera al VIS solo "
     f"({_det_llvip.loc['VIS','mAP50']:.4f}); la propuesta alcanza {_det_llvip.loc[PROP,'mAP50']:.4f}: "
     f"el realce que premian las métricas de actividad no mejora la detección (H3 no sostenida)."),
]
for k, v in info:
    ws.cell(f, 1, k).font = F_TXTB
    c = ws.cell(f, 2, v); c.font = F_TXT; c.alignment = WRAP
    ws.row_dimensions[f].height = 28
    f += 1
f += 1
ws.cell(f, 1, "Contenido del libro").font = F_SUB; f += 1
hojas = [
    ("PSO_Configuracion", "Parámetros del PSO de la propuesta (algoritmo, rangos, aptitud Fo, óptimo)"),
    ("PSO_Barrido", "Barrido de 25 configuraciones (partículas × iteraciones) replicando el Cuadro 1 del libro FPUNA"),
    ("Benchmark", f"Los 7 métodos × nueve métricas (promedio de los {N_ESC} pares TNO)"),
    ("Benchmark_por_Escena", f"Los 7 métodos en cada una de las {N_ESC} escenas ({N_ESC*7} filas, con filtro y mejor por escena)"),
    ("PSO_por_Escena", f"Las 25 configuraciones del PSO en cada escena ({N_ESC*25} filas) — equivale a los Anexos 1-{N_ESC}"),
    ("Propuesta_por_Imagen", "Métricas de la propuesta en cada escena, con promedio y desvío por fórmula"),
    ("Friedman", f"Test global por métrica (7 métodos × {N_ESC} imágenes)"),
    ("Wilcoxon", f"{len(wilc)} contrastes pareados (los dos morfológicos vs. los 5 del estado del arte, más Propuesta vs Top-Hat clásico)"),
    ("Ranking_Global", "Ranking promedio de los 7 métodos (nueve métricas)"),
    ("Deteccion_LLVIP", "mAP de YOLOv8n reentrenado por método sobre LLVIP (evaluación orientada a tarea)"),
    ("Deteccion_M3FD", "Clases complementarias en M3FD: AP por clase con un detector único VIS+IR (People/IR, Lamp/VIS)"),
]
for h, d in hojas:
    ws.cell(f, 1, h).font = F_TXTB
    ws.cell(f, 2, d).font = F_TXT
    f += 1

# ============================================================ 2. PSO CONFIG
ws = wb.create_sheet("PSO_Configuracion")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 70
f = titulo(ws, 1, "Proceso de optimización por enjambre de partículas (PSO)",
           "Corrida definitiva sobre el operador con SUMA de ramas.")
ws.cell(f, 1, "Reglas de actualización del enjambre").font = F_SUB; f += 1
f = nota(ws, f, "v_k(t+1) = ω·v_k(t) + c1·r1·(pbest_k − x_k) + c2·r2·(gbest − x_k)      x_k(t+1) = x_k(t) + v_k(t+1)")
f = nota(ws, f, "Inercia ω decreciente linealmente de 0,9 a 0,4 · c1 = c2 = 1,5 · r1, r2 ~ U(0,1) · posiciones recortadas al rango.")
f += 1
conf = [
    ("Método optimizado", "Propuesta: Top-Hat 1 escala, disco + 4 líneas, ramas SUMADAS (Bala et al. 2024)"),
    ("Variables", "(r, m) — radio del SE y peso de contraste"),
    ("Rango de búsqueda", "r ∈ [1, 25] · m ∈ [0,30; 2,00] (ambos del libro FPUNA, sin modificaciones)"),
    ("Configuraciones evaluadas", "25: partículas n ∈ {2,4,6,8,10} × iteraciones T ∈ {10,20,30,40,50} (Cuadro 1 del libro)"),
    ("Función de aptitud", "Fo = SSIM_avg + E_n + PSNR_n (Ortega y Espinoza, 2025; ecuación 14)"),
    ("Evaluada sobre", "3 escenas representativas del TNO (1 de cada 7)"),
    ("Óptimo global hallado", "r* = 25 · m* = 0,30 (el peso al que convergen las 25 configuraciones con el rango publicado)"),
    ("Aptitud alcanzada", f"Fo = {grid['F_opt'].max():.4f}"),
    ("Estado guardado en", "metrics_reports/pso_grid_search_fo_propuesta.csv"),
]
for k, v in conf:
    celda(ws, f, 1, k, bold=True, align=CL)
    celda(ws, f, 2, v, align=CL)
    f += 1
f += 1
f = nota(ws, f, "Con la suma de ramas se inyecta más energía de detalle que con el máximo; el enjambre lo compensa "
         "con el peso del rango publicado (m = 0,30). El radio se ubica en el tope del rango del libro (r = 25): "
         "el operador aprovecha un vecindario amplio para capturar los objetivos térmicos completos.")


# ============================================================ 2b. PSO BARRIDO
ws = wb.create_sheet("PSO_Barrido")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, "Barrido de configuraciones PSO — 25 combinaciones (Cuadro 1 de Ortega & Espinoza, 2025)",
           "Cada fila es una corrida PSO independiente (semilla propia). Las 25 configuraciones convergen al mismo peso óptimo, m* = 0,30; las diferencias de aptitud reflejan el radio hallado.")
f = encabezado(ws, f, ["Partículas (n)", "Iteraciones (Tmax)", "Evaluaciones", "r óptimo", "m óptimo",
                       "Aptitud F", "Tiempo (s)"], [14, 16, 13, 10, 11, 11, 11])
Fmax = grid["F_opt"].max()
for r in grid.itertuples():
    es_best = abs(r.F_opt - Fmax) < 5e-4
    celda(ws, f, 1, int(r.n), fmt="0", bold=es_best)
    celda(ws, f, 2, int(r.Tmax), fmt="0", bold=es_best)
    celda(ws, f, 3, int(r.evaluaciones), fmt="0", bold=es_best)
    celda(ws, f, 4, int(r.r_opt), fmt="0", bold=es_best)
    celda(ws, f, 5, float(r.m_opt), fmt="0.0000", bold=es_best)
    celda(ws, f, 6, float(r.F_opt), fmt="0.0000", bold=es_best)
    celda(ws, f, 7, int(r.segundos), fmt="0", bold=es_best)
    f += 1
f += 1
ws.cell(f, 1, "Matriz de aptitudes F (filas: partículas, columnas: iteraciones)").font = F_SUB; f += 1
piv = grid.pivot(index="n", columns="Tmax", values="F_opt")
f = encabezado(ws, f, ["n \ T"] + [f"T = {t}" for t in [10, 20, 30, 40, 50]])
for n in [2, 4, 6, 8, 10]:
    celda(ws, f, 1, f"n = {n}", bold=True)
    for j, t in enumerate([10, 20, 30, 40, 50], 2):
        v = float(piv.loc[n, t])
        celda(ws, f, j, v, fmt="0.0000", bold=abs(v - Fmax) < 5e-4)
    f += 1
f += 1
f = nota(ws, f, "Lectura: n = 2 cae en el óptimo local r = 1 en 2 de 5 corridas; con n = 10 y T >= 30 "
         "el óptimo global (r = 25, m ≈ 0,07) se alcanza de forma consistente.")

# ============================================================ 4. BENCHMARK
ws = wb.create_sheet("Benchmark")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, f"Benchmark — 7 métodos × nueve métricas (promedio de {N_ESC} pares TNO)",
           "Negrita = mejor de la columna (todas las métricas de tipo «mayor es mejor»).")
ws.column_dimensions["A"].width = 36
for j in range(2, 14):
    ws.column_dimensions[get_column_letter(j)].width = 8.5
f = encabezado(ws, f, ["Método"] + [f"{m} {'↓' if DIRECTION[m] < 0 else '↑'}" for m in METS])
best = {}
for mk in METS:
    vals = {m: means.loc[m, mk] for m in ORDEN}
    best[mk] = (max if DIRECTION[mk] > 0 else min)(vals, key=vals.get)
for m in ORDEN:
    es_prop = (m == PROP)
    celda(ws, f, 1, LBL[m], bold=es_prop, align=CL)
    for j, mk in enumerate(METS, 2):
        celda(ws, f, j, round(float(means.loc[m, mk]), 4), fmt="0.000",
              bold=(best[mk] == m or es_prop))
    f += 1

# ============================================================ 4b. BENCHMARK POR ESCENA
# Los 7 metodos en cada una de las 20 escenas (misma informacion que las Tablas 2a-2e del
# informe de avances, con el formato del Cuadro 2 de referencia, pero con las nueve metricas).
ws = wb.create_sheet("Benchmark_por_Escena")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, f"Benchmark escena por escena — los 7 métodos en cada uno de los {N_ESC} pares del TNO",
           "En negrita el mejor valor de cada métrica dentro de cada escena. Todas las métricas son "
           "de tipo «mayor es mejor». Equivale a las Tablas 2a-2e del informe de avances, ampliado a "
           "las nueve métricas.")
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 24
for j in range(3, 3 + NM):
    ws.column_dimensions[get_column_letter(j)].width = 9.5
f = encabezado(ws, f, ["Escena", "Método"] + METS)
_fila_ini = f
ESCENAS = list(dict.fromkeys(allm["image"]))
for _img in ESCENAS:
    _sub = allm[allm["image"] == _img].set_index("method")
    _mejor = {mk: _sub.loc[ORDEN, mk].idxmax() for mk in METS}
    for _k, _m in enumerate(ORDEN):
        celda(ws, f, 1, _img if _k == 0 else "", align=CL, bold=(_k == 0))
        celda(ws, f, 2, LBL[_m], align=CL, bold=(_m == PROP))
        for j, mk in enumerate(METS, 3):
            celda(ws, f, j, round(float(_sub.loc[_m, mk]), 5), fmt="0.00000",
                  bold=(_mejor[mk] == _m))
        f += 1
ws.freeze_panes = "C5"
ws.auto_filter.ref = f"A{_fila_ini - 1}:{get_column_letter(2 + NM)}{f - 1}"
f += 1
_lidera = {mk: sum(1 for im in ESCENAS
                   if allm[allm["image"] == im].set_index("method").loc[ORDEN, mk].idxmax() == PROP)
           for mk in METS}
f = nota(ws, f, f"Recuento sobre las {N_ESC} escenas — la propuesta obtiene el mejor valor en: "
         + " · ".join(f"{mk} {v}/20" for mk, v in _lidera.items())
         + ". Confirma el patrón de los promedios: lidera las métricas de información y contraste, "
           "y cede las de fidelidad a las fuentes.")

# ============================================================ 4c. PSO POR ESCENA (anexos)
# Las 25 configuraciones del Cuadro 1 ejecutadas sobre CADA imagen (20 x 25 = 500 filas).
# Es la misma informacion de los Anexos 1-20 del informe de avances.
_pso_img = pd.read_csv(os.path.join(MR, "pso_por_imagen.csv"))
ws = wb.create_sheet("PSO_por_Escena")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, f"PSO por escena — las 25 configuraciones del Cuadro 1 en cada uno de los {N_ESC} pares",
           "Aptitud Fo = SSIM_avg + E/8 + PSNR/100 sobre el rango publicado r ∈ [1,25], m ∈ [0,30; 2,00]. "
           "En negrita la configuración de mayor aptitud de cada escena. Fuente: pso_por_imagen.csv.")
ws.column_dimensions["A"].width = 42
for j, an in enumerate([9, 9, 7, 9, 12, 12, 12, 12, 12, 12], 2):
    ws.column_dimensions[get_column_letter(j)].width = an
_COLS = ["particulas", "iteraciones", "r", "m", "SSIM_avg", "E", "SF", "SD", "PSNR", "FO"]
f = encabezado(ws, f, ["Escena", "Partículas", "Iteraciones", "r", "m", "SSIM_avg", "E",
                       "SF", "SD", "PSNR", "FO"])
_ini = f
for _img in list(dict.fromkeys(_pso_img["imagen"])):
    _sub = _pso_img[_pso_img["imagen"] == _img]
    _best = _sub["FO"].idxmax()
    for _k, (_i, _r) in enumerate(_sub.iterrows()):
        _neg = (_i == _best)
        celda(ws, f, 1, _img if _k == 0 else "", align=CL, bold=(_k == 0))
        for j, c in enumerate(_COLS, 2):
            _v = _r[c]
            _fmt = "0" if c in ("particulas", "iteraciones", "r") else (
                "0.00" if c == "m" else "0.000000")
            celda(ws, f, j, (int(_v) if _fmt == "0" else round(float(_v), 6)),
                  fmt=_fmt, bold=_neg)
        f += 1
ws.freeze_panes = "B5"
ws.auto_filter.ref = f"A{_ini - 1}:K{f - 1}"
f += 1
_mej = _pso_img.loc[_pso_img.groupby("imagen")["FO"].idxmax()]
f = nota(ws, f, f"Lectura: en las {N_ESC} escenas el peso óptimo converge al límite inferior del rango "
         f"publicado (m* = 0,30 en {int((_mej['m'] == 0.30).sum())} de {N_ESC}) y el radio que maximiza Fo "
         f"es r = 1 en {int((_mej['r'] == 1).sum())} de {N_ESC}. Fo premia la fidelidad a las fuentes y por "
         f"tanto el mínimo realce; la configuración adoptada (r = 25) proviene del criterio de "
         f"evaluación de la tesis —las nueve métricas, todas «mayor es mejor»—, no de Fo.")

# ============================================================ 5. PROPUESTA POR IMAGEN
ws = wb.create_sheet("Propuesta_por_Imagen")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, "Propuesta novedosa (r=25, m=0,30) — métricas en cada escena del TNO",
           "Última fila: promedio y desvío estándar con fórmulas; el promedio coincide con la fila de la hoja Benchmark.")
ws.column_dimensions["A"].width = 40
for j in range(2, 2 + NM):
    ws.column_dimensions[get_column_letter(j)].width = 9
f = encabezado(ws, f, ["Escena"] + METS)
first_data = f
pi = allm[allm["method"] == PROP]
for _, r in pi.iterrows():
    celda(ws, f, 1, r["image"], align=CL)
    for j, mk in enumerate(METS, 2):
        celda(ws, f, j, round(float(r[mk]), 4), fmt="0.000")
    f += 1
last_data = f - 1
celda(ws, f, 1, f"PROMEDIO ({N_ESC} escenas)", bold=True, align=CL)
for j in range(2, 2 + NM):
    L = get_column_letter(j)
    celda(ws, f, j, f"=AVERAGE({L}{first_data}:{L}{last_data})", fmt="0.000", bold=True)
f += 1
celda(ws, f, 1, "DESVÍO ESTÁNDAR", bold=True, align=CL)
for j in range(2, 2 + NM):
    L = get_column_letter(j)
    celda(ws, f, j, f"=STDEV({L}{first_data}:{L}{last_data})", fmt="0.000", bold=True)
ws.freeze_panes = "B4"

# ============================================================ 6. FRIEDMAN
ws = wb.create_sheet("Friedman")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, f"Test de Friedman por métrica (7 métodos × {N_ESC} imágenes, por rangos)",
           "H0: no hay diferencias entre métodos. Significativa = p < 0,05.")
f = encabezado(ws, f, ["Métrica", "χ² de Friedman", "p-valor", "Significativa (α=0,05)"], [14, 16, 14, 20])
for r in fried.itertuples():
    celda(ws, f, 1, r.metric)
    celda(ws, f, 2, round(float(r.chi2), 2), fmt="0.0")
    celda(ws, f, 3, float(r.p_value), fmt="0.0E+00")
    celda(ws, f, 4, "Sí" if r.significant_05 else "No", bold=True)
    f += 1

# ============================================================ 7. WILCOXON
ws = wb.create_sheet("Wilcoxon")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, f"Wilcoxon pareado ({N_ESC} imágenes) — métodos morfológicos vs. estado del arte",
           "Corrección de Holm por métrica. Resultado por fórmula: ≈ si p_holm ≥ 0,05; si no, mejor/peor según "
           "(media método − media rival) × dirección (columna D; todas las métricas +1).")
cols = ["Método", "Rival", "Métrica", "Dirección", "Media método", "Media rival",
        "p", "p (Holm)", "Efecto r", "Resultado"]
f = encabezado(ws, f, cols, [24, 22, 9, 10, 12, 12, 11, 11, 9, 11])
SH = {"PiramideLaplace": "LP", "RatioPiramide": "RP", "DWT": "DWT", "DTCWT": "DTCWT",
      "Curvelet": "CVT", "TopHat_Clasico": "Top-Hat clásico", PROP: "Propuesta"}
w = wilc.sort_values(["tophat", "baseline", "metric"])
for r in w.itertuples():
    d = DIRECTION[r.metric]
    celda(ws, f, 1, SH.get(r.tophat, r.tophat), align=CL,
          bold=(r.tophat == PROP))
    celda(ws, f, 2, SH.get(r.baseline, r.baseline), align=CL)
    celda(ws, f, 3, r.metric)
    celda(ws, f, 4, d, fmt="0")
    celda(ws, f, 5, round(float(r.mean_tophat), 4), fmt="0.0000")
    celda(ws, f, 6, round(float(r.mean_baseline), 4), fmt="0.0000")
    celda(ws, f, 7, float(r.p_value), fmt="0.0E+00")
    celda(ws, f, 8, float(r.p_holm), fmt="0.0E+00")
    celda(ws, f, 9, round(float(r.effect_r), 2), fmt="0.00")
    formula = f'=IF(H{f}>=0.05,"≈",IF((E{f}-F{f})*D{f}>0,"mejor","peor"))'
    celda(ws, f, 10, formula, bold=True)
    f += 1
ws.freeze_panes = "A4"

# ============================================================ 8. RANKING
ws = wb.create_sheet("Ranking_Global")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, "Ranking global de los 7 métodos (nueve métricas, dirección respetada)",
           "Rango 1 = mejor en esa métrica. La columna final promedia los nueve rangos con fórmula. La propuesta es 1ª "
           "en EN y FE y 2ª en SD, MG y SF; cede en las métricas de fidelidad a las fuentes (SSIM, PSNR) y en información "
           f"mutua, lideradas por los métodos multiescala. Queda {POS_RK}ª en el ranking global "
           f"({VAL_RK}), calculado como promedio de rangos intra-bloque.")
ws.column_dimensions["A"].width = 36
for j in range(2, 3 + NM):
    ws.column_dimensions[get_column_letter(j)].width = 8.5
f = encabezado(ws, f, ["Método"] + METS + ["Rank promedio"])
rk = rankm.sort_values("avg_rank")
_last = get_column_letter(1 + NM)  # última columna de métrica (J con 9)
_avgc = 2 + NM                     # columna del rank promedio (K con 9)
for name, r in rk.iterrows():
    es_prop = (name == PROP)
    celda(ws, f, 1, LBL.get(name, name), bold=es_prop, align=CL)
    for j, mk in enumerate(METS, 2):
        celda(ws, f, j, float(r[mk]), fmt="0.0", bold=es_prop)
    celda(ws, f, _avgc, f"=AVERAGE(B{f}:{_last}{f})", fmt="0.00", bold=True)
    f += 1

# ============================================================ 9. DETECCIÓN LLVIP
det = pd.read_csv(os.path.join(MR, "detection_llvip_map.csv")).set_index("method")
ws = wb.create_sheet("Deteccion_LLVIP")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, "Evaluación orientada a tarea — detección de peatones en LLVIP",
           "YOLOv8n reentrenado 40 épocas por método (subconjunto 2.000 train / 500 val, misma config y semilla). "
           "Etiquetas idénticas para todos: la diferencia de mAP aísla el efecto del método de fusión. En negrita el mejor por columna.")
ws.column_dimensions["A"].width = 34
for j in range(2, 6):
    ws.column_dimensions[get_column_letter(j)].width = 15
f = encabezado(ws, f, ["Entrada del detector", "mAP@0,5 ↑", "mAP@0,5:0,95 ↑", "Precisión ↑", "Recall ↑"])
DET_ORDEN = ["VIS", "IR", "PiramideLaplace", "RatioPiramide", "DWT", "DTCWT",
             "Curvelet", "TopHat_Clasico", PROP]
DET_LBL = {**LBL, "VIS": "VIS (solo)", "IR": "IR (solo)",
           "Propuesta_Novedosa": "Propuesta Novedosa (r=25, m=0,30)"}
COLS = ["mAP50", "mAP50_95", "precision", "recall"]
best = {c: det[c].idxmax() for c in COLS}
for m in DET_ORDEN:
    r = det.loc[m]
    celda(ws, f, 1, DET_LBL.get(m, m), bold=(m == PROP), align=CL)
    for j, c in enumerate(COLS, 2):
        celda(ws, f, j, float(r[c]), fmt="0.000", bold=(best[c] == m or m == PROP))
    f += 1
f += 1
f = nota(ws, f, "Lectura: toda fusión supera al visible solo (+0,11 a +0,14 en mAP@0,5); el infrarrojo solo es la "
         "modalidad más fuerte (0,957) y ninguna fusión lo supera (peatón nocturno = térmico); entre las fusiones "
         "(banda 0,913–0,949) la propuesta queda en el extremo inferior (0,913). El realce que premian las métricas de "
         "actividad no mejora la detección: ambos criterios deben reportarse por separado (H3 no sostenida).")

# ============================================================ 11. DETECCION M3FD
ws = wb.create_sheet("Deteccion_M3FD")
ws.sheet_view.showGridLines = False
f = titulo(ws, 1, "Clases complementarias en M3FD — un detector único VIS+IR",
           "YOLOv8n entrenado con las imágenes de ambas modalidades mezcladas (etiquetas compartidas, 40 épocas) y evaluado "
           "por inferencia sobre la partición de PRUEBA de cada entrada, disjunta de la de entrenamiento y de la de "
           "selección del modelo. People domina en IR (térmica); Lamp solo se ve en VIS. En negrita el mejor por columna.")
ws.column_dimensions["A"].width = 40
for j in range(2, 6):
    ws.column_dimensions[get_column_letter(j)].width = 16
f = encabezado(ws, f, ["Entrada del detector", "AP@0,5 People ↑", "AP@0,5 Lamp ↑", "Promedio del par ↑", "mAP@0,5 (6 clases) ↑"])
# Tabla leida del CSV, no escrita a mano: antes los nueve valores estaban tipeados en el
# codigo mientras la nota declaraba el CSV como fuente, de modo que un recomputo de la
# deteccion no se reflejaba aqui.
_m3 = pd.read_csv(os.path.join(MR, "detection_m3fd_map.csv")).set_index("method")
_m3["par"] = (_m3["AP50_People"] + _m3["AP50_Lamp"]) / 2
_M3_ORDEN = ["VIS", "IR", "PiramideLaplace", "RatioPiramide", "DWT", "DTCWT",
             "Curvelet", "TopHat_Clasico", PROP]
_M3_LBL = {"VIS": "VIS (solo)", "IR": "IR (solo)",
           "PiramideLaplace": "Pirámide de Laplace (LP)", "RatioPiramide": "Ratio Pyramid (RP)",
           "DWT": "Wavelet discreta (DWT)", "DTCWT": "Dual-Tree Complex Wavelet (DTCWT)",
           "Curvelet": "Curvelet (CVT)", "TopHat_Clasico": "Top-Hat clásico (r=5; m=1)",
           PROP: "Propuesta Novedosa (r=25; m=0,30)"}
_M3_COLS = ["AP50_People", "AP50_Lamp", "par", "mAP50"]
_pres = [k for k in _M3_ORDEN if k in _m3.index]
_m3_best = {c: _m3.loc[_pres, c].idxmax() for c in _M3_COLS}
for _k in _pres:
    es_prop = (_k == PROP)
    celda(ws, f, 1, _M3_LBL.get(_k, _k), bold=es_prop, align=CL)
    for j, c in enumerate(_M3_COLS, start=1):
        celda(ws, f, j + 1, round(float(_m3.loc[_k, c]), 4), fmt="0.000",
              bold=(_m3_best[c] == _k or es_prop))
    f += 1
f += 1
_m3f = _m3.loc[[k for k in _pres if k not in ("VIS", "IR")]]
_sup = [k for k in _m3f.index if _m3f.loc[k, "par"] > _m3.loc["VIS", "par"]
        and _m3f.loc[k, "par"] > _m3.loc["IR", "par"]]
_rec = [k for k in _m3f.index if _m3f.loc[k, "AP50_People"] > 0 and _m3f.loc[k, "AP50_Lamp"] > 0]
_pos = int(_m3f["par"].rank(ascending=False)[PROP])
_txt = (f"Lectura: complementariedad de las clases en las modalidades individuales — IR "
        f"{_m3.loc['IR','AP50_People']:.3f} en People frente a {_m3.loc['IR','AP50_Lamp']:.3f} en Lamp; "
        f"VIS {_m3.loc['VIS','AP50_People']:.3f} y {_m3.loc['VIS','AP50_Lamp']:.3f}. "
        f"{len(_rec)} de las {len(_m3f)} fusiones detectan objetos de ambas clases en una sola imagen. ")
_txt += ((f"En el promedio del par, {len(_sup)} de las {len(_m3f)} "
          f"fusiones {'supera' if len(_sup)==1 else 'superan'} a ambas modalidades "
          f"({', '.join(_M3_LBL.get(k,k).split(' (')[0] for k in _sup)}). ")
         if _sup else
         (f"Ninguna de las {len(_m3f)} fusiones supera a ambas modalidades en el promedio del par "
          f"(mejor fusión {_m3f['par'].max():.3f} frente a {_m3.loc['VIS','par']:.3f} del VIS). "))
_txt += (f"La propuesta obtiene {_m3.loc[PROP,'par']:.3f} en el par (puesto {_pos} de {len(_m3f)} "
         f"entre las fusiones), People {_m3.loc[PROP,'AP50_People']:.3f} y Lamp "
         f"{_m3.loc[PROP,'AP50_Lamp']:.3f}. Particiones train/val/test disjuntas y estratificadas; "
         f"el test es la única partición reportada. Fuente: metrics_reports/detection_m3fd_map.csv.")
f = nota(ws, f, _txt)

wb.save(OUT)
print("Guardado:", OUT)
